import os
import time
import gc
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from concurrent.futures import ProcessPoolExecutor, as_completed
from functools import lru_cache

OLD_VALUES = {
    "白極Marsh": "白極",
    "氷見忘れた": "氷見",
    "彰良": "Chitose彰良",
    "マイキ": "マイキP",
}
TARGET_COLUMN_NAME = "author"
PARTIAL_REPLACE = False
BATCH_SIZE = 100
MAX_WORKERS = 12


def text_scan(file_path: str) -> bool:
    try:
        import zipfile

        with zipfile.ZipFile(file_path, "r") as z:
            if "xl/sharedStrings.xml" in z.namelist():
                content = z.read("xl/sharedStrings.xml").decode(
                    "utf-8", errors="ignore"
                )
                return any(old_val in content for old_val in OLD_VALUES.keys())
        return False
    except:
        return True


def process_excel_file(file_path: str) -> tuple[str, str]:
    """处理单个文件"""
    try:
        if not text_scan(file_path):
            return file_path, "无需修改"

        wb = load_workbook(file_path)
        modified_count = 0
        old_values_set = set(OLD_VALUES.keys())

        for ws in wb.worksheets:
            col_idx = None

            if TARGET_COLUMN_NAME:
                for col in ws.iter_cols(min_row=1, max_row=1):
                    if col[0].value == TARGET_COLUMN_NAME:
                        col_idx = col[0].column
                        break
                if col_idx is None:
                    continue

            if col_idx:
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        if isinstance(c, MergedCell) or not isinstance(c.value, str):
                            continue

                        if PARTIAL_REPLACE:
                            original = c.value
                            for old_val, new_val in OLD_VALUES.items():
                                if old_val in c.value:
                                    c.value = c.value.replace(old_val, new_val)
                            if c.value != original:
                                modified_count += 1
                        else:
                            if c.value in old_values_set:
                                c.value = OLD_VALUES[c.value]
                                modified_count += 1

        if modified_count > 0:
            wb.save(file_path)
            wb.close()
            return file_path, f"修改{modified_count}处"

        wb.close()
        return file_path, "无需修改"

    except Exception as e:
        return file_path, f"出错: {e}"


def process_batch(batch: list[str]) -> list[tuple[str, str]]:
    """处理一批文件"""
    results = []
    for fp in batch:
        results.append(process_excel_file(fp))
    gc.collect()
    return results


def main():
    start_time = time.time()
    root_folder = os.path.abspath(os.path.dirname(__file__))

    filepaths = [
        os.path.join(dirpath, filename)
        for dirpath, _, filenames in os.walk(root_folder)
        for filename in filenames
        if filename.lower().endswith(".xlsx") and not filename.startswith("~$")
    ]

    total = len(filepaths)
    if not total:
        print("未找到任何 .xlsx 文件。")
        return

    print(f"找到 {total} 个文件，分 {(total // BATCH_SIZE) + 1} 批处理...\n")

    modified_files = []
    error_files = []
    processed = 0

    for i in range(0, total, BATCH_SIZE):
        batch = filepaths[i : i + BATCH_SIZE]

        with ProcessPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [executor.submit(process_excel_file, fp) for fp in batch]

            for future in as_completed(futures):
                file_path, message = future.result()
                processed += 1

                if "出错" in message:
                    error_files.append((file_path, message))
                elif message != "无需修改":
                    modified_files.append((file_path, message))

        progress = processed / total * 100
        print(
            f"\r进度: {processed}/{total} ({progress:.1f}%) | 已修改: {len(modified_files)}",
            end="",
            flush=True,
        )

        gc.collect()

    # 结果汇总
    print(f"\n\n{'='*50}")
    print(f"处理完成！耗时: {time.time() - start_time:.1f} 秒")
    print(f"总文件: {total} | 修改: {len(modified_files)} | 错误: {len(error_files)}")

    if modified_files:
        print(f"\n修改的文件:")
        for fp, msg in modified_files[:20]:  # 只显示前20个
            print(f"  {os.path.basename(fp)}: {msg}")
        if len(modified_files) > 20:
            print(f"  ... 还有 {len(modified_files) - 20} 个文件")

    if error_files:
        print(f"\n出错的文件:")
        for fp, msg in error_files[:10]:
            print(f"  {os.path.basename(fp)}: {msg}")


if __name__ == "__main__":
    main()

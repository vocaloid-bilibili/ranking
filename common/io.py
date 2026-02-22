# common/io.py
"""文件读写工具"""

import pandas as pd
from pathlib import Path
from typing import Optional, List, Dict, Union

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment

from common.logger import logger
from common.formatters import clean_excel_chars, format_aid, COLUMN_FORMATTERS


# ==================== DataFrame 预处理 ====================


def clean_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """清洗DataFrame中的非法字符"""
    df = df.copy()
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].apply(clean_excel_chars)
    return df


def prepare_dataframe(
    df: pd.DataFrame, usecols: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    准备DataFrame以供保存

    Args:
        df: 原始DataFrame
        usecols: 需要保存的列
    """
    if usecols:
        df = df[[c for c in usecols if c in df.columns]].copy()
    else:
        df = df.copy()

    # 格式化aid列
    if "aid" in df.columns:
        df["aid"] = df["aid"].apply(format_aid)

    # 应用列格式化器（tid → tname 等）
    for col, formatter in COLUMN_FORMATTERS.items():
        if col in df.columns:
            df[col] = df[col].apply(formatter)

    # 清洗非法字符
    return clean_dataframe_for_excel(df)


# ==================== Excel 格式设置 ====================


class ExcelStyler:
    """Excel样式设置器"""

    # 需要设置为文本格式的列
    TEXT_COLUMNS = ["pubdate", "aid"]

    @staticmethod
    def apply_text_format(worksheet, df: pd.DataFrame, columns: List[str] = None):
        """将指定列设置为文本格式"""
        columns = columns or ExcelStyler.TEXT_COLUMNS

        for col_name in columns:
            if col_name not in df.columns:
                continue

            col_idx = df.columns.get_loc(col_name)
            if not isinstance(col_idx, int):
                continue

            letter = get_column_letter(col_idx + 1)
            for cell in worksheet[letter]:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")

    @staticmethod
    def apply_row_colors(worksheet, df: pd.DataFrame, row_styles: Dict[int, str]):
        """设置行背景色"""
        for idx, color in row_styles.items():
            if idx not in df.index:
                logger.warning(f"索引 {idx} 不在DataFrame中，跳过样式设置")
                continue

            row_num = df.index.get_loc(idx) + 2  # Excel行号从1开始，加上标题行
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            for cell in worksheet[row_num]:
                cell.fill = fill


# ==================== 读写函数 ====================


def save_excel(
    df: pd.DataFrame,
    path: Union[str, Path],
    usecols: Optional[List[str]] = None,
    row_styles: Optional[Dict[int, str]] = None,
    text_columns: Optional[List[str]] = None,
):
    """
    保存DataFrame到Excel文件

    Args:
        df: 要保存的DataFrame
        path: 保存路径
        usecols: 需要保存的列
        row_styles: 行样式 {行索引: 颜色hex}
        text_columns: 需要设为文本格式的列
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    df = prepare_dataframe(df, usecols)

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]

            # 应用文本格式
            ExcelStyler.apply_text_format(ws, df, text_columns)

            # 应用行颜色
            if row_styles:
                ExcelStyler.apply_row_colors(ws, df, row_styles)

        logger.info(f"保存完成: {path}")

    except Exception as e:
        logger.warning(f"Excel保存失败: {e}")
        _save_csv_backup(df, path)


def _save_csv_backup(df: pd.DataFrame, original_path: Path):
    """保存CSV备份"""
    backup = original_path.with_suffix(".csv")
    df.to_csv(backup, index=False, encoding="utf-8-sig")
    logger.info(f"已备份至: {backup}")


def load_excel(
    path: Union[str, Path],
    usecols: Optional[List[str]] = None,
    dtype: Optional[Dict[str, type]] = None,
) -> pd.DataFrame:
    """
    读取Excel文件

    Args:
        path: 文件路径
        usecols: 需要读取的列
        dtype: 列数据类型
    """
    path = Path(path)
    if not path.exists():
        logger.warning(f"文件不存在: {path}")
        return pd.DataFrame()

    try:
        return pd.read_excel(path, usecols=usecols, dtype=dtype)
    except Exception as e:
        logger.error(f"读取失败: {path}, 错误: {e}")
        return pd.DataFrame()


def load_csv(
    path: Union[str, Path],
    usecols: Optional[List[str]] = None,
    encoding: str = "utf-8-sig",
) -> pd.DataFrame:
    """
    读取CSV文件

    Args:
        path: 文件路径
        usecols: 需要读取的列
        encoding: 编码
    """
    path = Path(path)
    if not path.exists():
        logger.warning(f"文件不存在: {path}")
        return pd.DataFrame()

    try:
        return pd.read_csv(path, usecols=usecols, encoding=encoding)
    except Exception as e:
        logger.error(f"读取失败: {path}, 错误: {e}")
        return pd.DataFrame()


def ensure_dir(path: Union[str, Path]) -> Path:
    """确保目录存在"""
    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)
    return path


def get_latest_file(directory: Path, pattern: str = "*.xlsx") -> Optional[Path]:
    """获取目录中最新的文件"""
    files = list(directory.glob(pattern))
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)
